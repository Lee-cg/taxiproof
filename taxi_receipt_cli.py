#!/usr/bin/env python3
"""
택시비 증빙 자료 생성 프로그램 (CLI 버전)
- 여러 택시비 이미지를 가로로 이어붙여 하나의 이미지 생성
- 각 이미지에서 결제 금액과 일시를 추출하여 엑셀 파일 생성
"""

import os
import re
import sys
import argparse
from datetime import datetime
from PIL import Image
import easyocr
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class TaxiReceiptProcessor:
    def __init__(self):
        self.reader = None

    def init_ocr(self):
        """OCR 리더 초기화"""
        if self.reader is None:
            print("OCR 엔진 초기화 중... (최초 실행시 모델 다운로드)")
            self.reader = easyocr.Reader(['ko', 'en'], gpu=False)
        return self.reader

    def merge_images_horizontal(self, image_paths, output_path):
        """이미지들을 가로로 이어붙이기"""
        images = [Image.open(p) for p in image_paths]

        # 가장 높은 이미지 기준으로 높이 통일
        max_height = max(img.height for img in images)

        # 비율 유지하며 리사이즈
        resized = []
        total_width = 0
        for img in images:
            ratio = max_height / img.height
            new_width = int(img.width * ratio)
            resized_img = img.resize((new_width, max_height), Image.Resampling.LANCZOS)
            resized.append(resized_img)
            total_width += new_width

        # 새 이미지 생성 및 붙이기
        merged = Image.new('RGB', (total_width, max_height), (255, 255, 255))
        x_offset = 0
        for img in resized:
            merged.paste(img, (x_offset, 0))
            x_offset += img.width

        merged.save(output_path, quality=95)

        # 리소스 정리
        for img in images:
            img.close()
        for img in resized:
            img.close()

        return output_path

    def extract_receipt_info(self, image_path):
        """이미지에서 결제 금액, 일시, 운행종료시간 추출"""
        reader = self.init_ocr()
        result = reader.readtext(image_path)

        # 전체 텍스트 추출
        full_text = ' '.join([item[1] for item in result])

        amount = self.extract_amount(full_text)
        date_time = self.extract_datetime(full_text)
        end_time = self.extract_end_time(full_text)

        return amount, date_time, end_time

    def extract_end_time(self, text):
        """운행 시간에서 종료 시간 추출 (22:10 - 22:27 또는 22.52 23.06 -> 종료시간)"""
        # 22:10 - 22:27 또는 22.52 23.06 형식 (하이픈 없을 수도 있음, :가 .으로 인식될 수 있음)
        pattern = r'\d{1,2}[.:]\d{2}\s*[-~]?\s*(\d{1,2})([.:])(\d{2})'
        match = re.search(pattern, text)
        if match:
            hour = match.group(1)
            minute = match.group(3)
            return f"{hour}:{minute}"  # 항상 : 형식으로 반환
        return None

    def extract_amount(self, text):
        """결제 금액 추출"""
        patterns = [
            r'결제\s*금액\s*[:\s]*([0-9,]+)\s*원?',
            r'총\s*금액\s*[:\s]*([0-9,]+)\s*원?',
            r'합\s*계\s*[:\s]*([0-9,]+)\s*원?',
            r'카드\s*결제\s*[:\s]*([0-9,]+)\s*원?',
            r'승인\s*금액\s*[:\s]*([0-9,]+)\s*원?',
            r'([0-9]{1,3}(?:,?[0-9]{3})+)\s*원',
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                amount_str = match.group(1).replace(',', '')
                try:
                    return int(amount_str)
                except ValueError:
                    continue

        return None

    def correct_ocr_date(self, date_str):
        """OCR 날짜 인식 오류 보정 (7->1 오인식 등)"""
        if not date_str:
            return date_str

        # YY.MM.DD 형식 파싱
        match = re.match(r'(\d{2})\.(\d{2})\.(\d{2})', date_str)
        if match:
            year, month, day = match.groups()
            month = int(month)
            day = int(day)

            # 월이 12 초과면 7->1 보정 시도
            if month > 12:
                corrected = str(month).replace('7', '1')
                if int(corrected) <= 12:
                    month = int(corrected)

            # 일이 31 초과면 7->1 보정 시도
            if day > 31:
                corrected = str(day).replace('7', '1')
                if int(corrected) <= 31:
                    day = int(corrected)

            return f"{year}.{month:02d}.{day:02d}"

        return date_str

    def extract_datetime(self, text):
        """결제 일시 추출"""
        patterns = [
            # 26.01.12. 22.27 또는 26.01.12 22:27 형식 (OCR이 :를 .으로 인식할 수 있음)
            r'(\d{2}\.\d{2}\.\d{2})\.?\s*(\d{1,2})[.:](\d{2})',
            # 2024-01-15 14:30:25 또는 2024.01.15 14:30:25
            r'(\d{4}[-./]\d{1,2}[-./]\d{1,2})\.?\s*(\d{1,2})[.:](\d{2})',
            # 24-01-15 14:30 또는 24.01.15 14:30
            r'(\d{2}[-./]\d{1,2}[-./]\d{1,2})\.?\s*(\d{1,2})[.:](\d{2})',
            # 날짜만
            r'(\d{4}[-./]\d{1,2}[-./]\d{1,2})()()',
            r'(\d{2}\.\d{2}\.\d{2})()()',
        ]

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                date_part = match.group(1)
                hour = match.group(2) if match.group(2) else None
                minute = match.group(3) if match.group(3) else None

                # 날짜 보정
                date_part = self.correct_ocr_date(date_part)

                if hour and minute:
                    return f"{date_part} {hour}:{minute}"
                return date_part

        return None

    def parse_datetime(self, date_time_str):
        """날짜/시간 문자열을 분리"""
        if not date_time_str:
            return None, None

        # "26.01.12 22:27" 형식 파싱
        parts = date_time_str.strip().split()
        if len(parts) >= 2:
            date_part = parts[0]  # 26.01.12
            time_part = parts[1]  # 22:27

            # 날짜를 2026.01.12 형식으로 변환
            if len(date_part.split('.')[0]) == 2:
                date_part = '20' + date_part  # 26.01.12 -> 2026.01.12

            return date_part, time_part
        elif len(parts) == 1:
            # 날짜만 있는 경우
            date_part = parts[0]
            if len(date_part.split('.')[0]) == 2:
                date_part = '20' + date_part
            return date_part, None

        return None, None

    def create_excel(self, data, output_path, user_name="레논", usage="야근"):
        """엑셀 파일 생성 (개인경비관리시트 양식 - 26_01.xlsx 스타일)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "야근택시비"

        # 스타일 정의
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(bold=True, size=9)
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font_white = Font(bold=True, size=9, color="FFFFFF")
        data_font = Font(size=9)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # 행 1: 제목 (병합)
        ws.merge_cells('A1:G1')
        cell = ws.cell(row=1, column=1, value="개인경비관리시트(2026)")
        cell.font = Font(bold=True, size=10)
        cell.alignment = left_align
        cell.border = thin_border

        # 행 2: 안내문
        ws.merge_cells('A2:G2')
        cell = ws.cell(row=2, column=1, value="개인카드로 결제 후 정산하시고자 하는 경우 개인경비관리시트를 작성하여 파일첨부, 원본영수증을 제출해주셔야 정산 가능합니다.")
        cell.font = Font(size=11)
        cell.alignment = left_align
        cell.border = thin_border

        # 행 3: 안내문
        ws.merge_cells('A3:G3')
        cell = ws.cell(row=3, column=1, value="접대비는 원칙적으로 개인카드 정산이 불가합니다.")
        cell.font = Font(size=9, color="FF0000")
        cell.alignment = left_align
        cell.border = thin_border

        # 행 4: 빈 행

        # 행 5: 헤더
        headers = ["계정", "승인일자", "승인시간", "사용자", "사용내역", "금액", "비고"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 데이터 입력 (6행부터)
        total_amount = 0
        data_start_row = 6
        for row_idx, (filename, amount, date_time, end_time) in enumerate(data, data_start_row):
            # 날짜만 분리 (시간은 end_time 사용)
            date_str, _ = self.parse_datetime(date_time)

            # A: 계정
            cell = ws.cell(row=row_idx, column=1, value="여비교통비(시내교통)")
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # B: 승인일자
            cell = ws.cell(row=row_idx, column=2, value=date_str if date_str else "인식 실패")
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # C: 승인시간 (운행 종료 시간, 10:27:00 PM 형식)
            if end_time:
                try:
                    from datetime import datetime as dt
                    time_obj = dt.strptime(end_time, "%H:%M")
                    cell = ws.cell(row=row_idx, column=3, value=time_obj)
                    cell.number_format = "h:mm:ss AM/PM"
                except:
                    cell = ws.cell(row=row_idx, column=3, value=end_time)
            else:
                cell = ws.cell(row=row_idx, column=3, value="인식 실패")
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # D: 사용자
            cell = ws.cell(row=row_idx, column=4, value=user_name)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # E: 사용내역
            cell = ws.cell(row=row_idx, column=5, value=usage)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # F: 금액
            if amount is not None:
                cell = ws.cell(row=row_idx, column=6, value=amount)
                cell.number_format = '#,##0'
                total_amount += amount
            else:
                cell = ws.cell(row=row_idx, column=6, value="인식 실패")
            cell.font = data_font
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border

            # G: 비고
            cell = ws.cell(row=row_idx, column=7, value=None)
            cell.border = thin_border

        # 빈 행 추가
        data_end_row = data_start_row + len(data) - 1
        summary_start = data_end_row + 3

        # 금액 열 범위 (F6:F{data_end_row})
        sum_range = f"F{data_start_row}:F{data_end_row}"

        # 계정별 금액합계 헤더
        cell = ws.cell(row=summary_start, column=1, value="계정")
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        cell = ws.cell(row=summary_start, column=2, value="금액합계")
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # 계정별 합계 (여비교통비만) - SUM 함수 사용
        cell = ws.cell(row=summary_start + 1, column=1, value="여비교통비(시내교통)")
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border

        cell = ws.cell(row=summary_start + 1, column=2, value=f"=SUM({sum_range})")
        cell.font = data_font
        cell.number_format = '#,##0'
        cell.border = thin_border

        # 합계 행 - SUM 함수 사용
        cell = ws.cell(row=summary_start + 2, column=1, value="합 계")
        cell.font = Font(bold=True, size=9)
        cell.alignment = center_align
        cell.border = thin_border

        cell = ws.cell(row=summary_start + 2, column=2, value=f"=SUM({sum_range})")
        cell.font = Font(bold=True, size=9)
        cell.number_format = '#,##0'
        cell.border = thin_border

        # 열 너비 조정
        ws.column_dimensions['A'].width = 20.5
        ws.column_dimensions['B'].width = 16.35
        ws.column_dimensions['C'].width = 17.67
        ws.column_dimensions['D'].width = 15.67
        ws.column_dimensions['E'].width = 17.17
        ws.column_dimensions['F'].width = 14.5
        ws.column_dimensions['G'].width = 41.85

        wb.save(output_path)

    def process(self, image_paths, output_dir, prefix="택시비_증빙", user_name="레논", usage="야근"):
        """메인 처리 로직"""
        if not image_paths:
            print("오류: 처리할 이미지가 없습니다.")
            return False

        if not os.path.isdir(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            # 1. 이미지 병합
            print(f"\n[1/3] 이미지 병합 중... ({len(image_paths)}개)")
            merged_path = os.path.join(output_dir, f"{prefix}_{timestamp}.png")
            self.merge_images_horizontal(image_paths, merged_path)
            print(f"      완료: {merged_path}")

            # 2. OCR로 정보 추출
            print(f"\n[2/3] OCR 처리 중...")
            extracted_data = []
            for i, path in enumerate(image_paths):
                print(f"      ({i+1}/{len(image_paths)}) {os.path.basename(path)}")
                amount, date_time, end_time = self.extract_receipt_info(path)
                extracted_data.append((os.path.basename(path), amount, date_time, end_time))

                # 결과 출력
                amount_str = f"{amount:,}원" if amount else "인식 실패"
                date_str = date_time if date_time else "인식 실패"
                time_str = end_time if end_time else "인식 실패"
                print(f"          금액: {amount_str}, 일시: {date_str}, 종료시간: {time_str}")

            # 3. 엑셀 파일 생성
            print(f"\n[3/3] 엑셀 파일 생성 중...")
            excel_path = os.path.join(output_dir, f"{prefix}_{timestamp}.xlsx")
            self.create_excel(extracted_data, excel_path, user_name, usage)
            print(f"      완료: {excel_path}")

            # 결과 요약
            total = sum(d[1] for d in extracted_data if d[1] is not None)
            print(f"\n{'='*50}")
            print(f"처리 완료!")
            print(f"  - 이미지 파일: {merged_path}")
            print(f"  - 엑셀 파일: {excel_path}")
            print(f"  - 총 금액: {total:,}원")
            print(f"{'='*50}")

            return True

        except Exception as e:
            print(f"\n오류 발생: {str(e)}")
            return False


def get_image_files(paths):
    """경로에서 이미지 파일 목록 추출"""
    image_extensions = {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp'}
    image_files = []

    for path in paths:
        if os.path.isfile(path):
            if os.path.splitext(path)[1].lower() in image_extensions:
                image_files.append(path)
        elif os.path.isdir(path):
            for filename in sorted(os.listdir(path)):
                filepath = os.path.join(path, filename)
                if os.path.isfile(filepath):
                    if os.path.splitext(filename)[1].lower() in image_extensions:
                        image_files.append(filepath)

    return image_files


def main():
    parser = argparse.ArgumentParser(
        description='택시비 증빙 자료 생성기 (CLI)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  python taxiproof.py image1.jpg image2.jpg image3.jpg
  python taxiproof.py ./receipts/
  python taxiproof.py *.jpg -o ~/Desktop -p 1월_택시비
  python taxiproof.py *.jpg -u 홍길동 -d 야근
        """
    )

    parser.add_argument('images', nargs='+', help='이미지 파일 또는 디렉토리')
    parser.add_argument('-o', '--output', default=os.path.expanduser('~/Desktop'),
                        help='출력 디렉토리 (기본값: ~/Desktop)')
    parser.add_argument('-p', '--prefix', default='택시비_증빙',
                        help='출력 파일명 접두사 (기본값: 택시비_증빙)')
    parser.add_argument('-u', '--user', default='레논',
                        help='사용자 이름 (기본값: 레논)')
    parser.add_argument('-d', '--desc', default='야근',
                        help='사용내역 (기본값: 야근)')

    args = parser.parse_args()

    # 이미지 파일 수집
    image_files = get_image_files(args.images)

    if not image_files:
        print("오류: 처리할 이미지 파일을 찾을 수 없습니다.")
        sys.exit(1)

    print(f"택시비 증빙 자료 생성기")
    print(f"{'='*50}")
    print(f"입력 이미지: {len(image_files)}개")
    for f in image_files:
        print(f"  - {os.path.basename(f)}")
    print(f"출력 위치: {args.output}")
    print(f"파일명 접두사: {args.prefix}")
    print(f"사용자: {args.user}")
    print(f"사용내역: {args.desc}")

    # 처리 실행
    processor = TaxiReceiptProcessor()
    success = processor.process(image_files, args.output, args.prefix, args.user, args.desc)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
